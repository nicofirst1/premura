"""Personal lab-history spreadsheet parser (wide xlsx -> long Measurement rows).

The source files are wide "test x date" grids exported from a lab portal: one
sheet per language (``Italian`` authoritative, ``German`` a test-name-only
translation), header row ``["test_name", "unit", "reference_range", <date>,
...]``, one data row per test with a value in each date column that was
actually collected. This parser unpivots that grid into one ``Measurement``
per (test, date) cell that has a value.

Distinct from ``lab_pdf`` (source_kind ``lab_pdf``, PDF-only, autodiscovers
``*.pdf``): this parser's source_kind is ``labsheet`` and it autodiscovers
``*.xlsx``. Both funnel into the same ``lab:*`` ontology namespace via
``suggest_metric`` so overlapping tests resolve to the same metric_id.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]

from .base import IngestBatch, Measurement, SourceDescriptor
from .lookup import suggest_metric

SOURCE_KIND = "labsheet"
SOURCE_ID = "labsheet:personal"
_PRIMARY_SHEET = "Italian"
_TRANSLATION_SHEET = "German"
_HEADER_FIXED = ("test_name", "unit", "reference_range")


@dataclass(slots=True)
class LabXlsxParser:
    source_kind: str = SOURCE_KIND
    language_hint: str | None = "it"

    def declares_metrics(self) -> list[str]:
        # suggest_metric (decision-tree step 1) is the only mapping step this
        # parser uses, and it can return any ontology metric_id -- most rows
        # land under lab:*, but a handful of tests (e.g. "Glicemia") already
        # have a bare cross-vendor alias (blood_glucose) instead. Declare the
        # full ontology so declares_metrics always matches what parse() may
        # emit, whichever namespace suggest_metric resolves into.
        from .lookup import metric_ids

        return sorted(metric_ids())

    def parse(self, path: Path) -> IngestBatch:
        wb = openpyxl.load_workbook(path, data_only=True)
        if _PRIMARY_SHEET not in wb.sheetnames:
            raise ValueError(f"{path.name}: missing required {_PRIMARY_SHEET!r} sheet")
        sheet = wb[_PRIMARY_SHEET]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"{path.name}: {_PRIMARY_SHEET} sheet is empty")

        header = rows[0]
        _validate_header(header, path)
        date_cols = _parse_date_headers(header[3:], offset=3, path=path)

        translations = _load_translations(wb)

        result = IngestBatch(
            source_kind=SOURCE_KIND,
            declared_metrics=self.declares_metrics(),
            source_descriptors={
                SOURCE_ID: SourceDescriptor(source_id=SOURCE_ID, source_kind=SOURCE_KIND)
            },
        ).attach_source_artifact(path)

        seen_uuids: set[str] = set()
        for row in rows[1:]:
            test_name = row[0]
            if not test_name or not str(test_name).strip():
                continue
            test_name = str(test_name).strip()
            raw_unit = row[1]
            reference_range = row[2]

            metric_id = suggest_metric(test_name)
            if metric_id is None:
                vendor_field = f"vendor:{SOURCE_KIND}:{_slugify(test_name)}"
                if vendor_field not in result.unmapped_metrics:
                    result.unmapped_metrics.append(vendor_field)
                continue

            for col_index, ts_utc in date_cols:
                raw_value = row[col_index] if col_index < len(row) else None
                if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
                    continue

                value_num, value_text = _parse_cell_value(raw_value)
                if value_num is None and value_text is None:
                    continue

                source_uuid = _measurement_uuid(metric_id, ts_utc, value_num, value_text)
                if source_uuid in seen_uuids:
                    # Same (metric, date, value) already emitted -- the
                    # Stools/Blood-Urine file overlap collapses here even
                    # before the warehouse dedupe_key UNIQUE constraint.
                    continue
                seen_uuids.add(source_uuid)

                payload: dict[str, Any] = {
                    "original_test_name": test_name,
                    "original_unit": raw_unit,
                    "reference_range": reference_range,
                    "sheet_language": "it",
                    "translation_de": translations.get(test_name),
                    "source_file": path.name,
                }

                result.measurements.append(
                    Measurement(
                        ts_utc=ts_utc,
                        metric_id=metric_id,
                        unit=_normalize_unit(raw_unit),
                        source_id=SOURCE_ID,
                        source_kind=SOURCE_KIND,
                        value_num=value_num,
                        value_text=value_text,
                        source_uuid=source_uuid,
                        raw_payload=payload,
                    )
                )

        result.validate()
        return result


def _validate_header(header: tuple[Any, ...], path: Path) -> None:
    got = tuple(str(h).strip().lower() if h else h for h in header[:3])
    if got != _HEADER_FIXED:
        raise ValueError(f"{path.name}: expected header {_HEADER_FIXED!r}, got {header[:3]!r}")


def _parse_date_headers(
    date_header_cells: tuple[Any, ...], *, offset: int, path: Path
) -> list[tuple[int, datetime]]:
    date_cols: list[tuple[int, datetime]] = []
    for i, cell in enumerate(date_header_cells):
        if cell is None:
            continue
        if not isinstance(cell, datetime):
            raise ValueError(
                f"{path.name}: date column header at position {offset + i} is not a date: {cell!r}"
            )
        # Bare-date posture: midnight, naive, no invented timezone.
        date_cols.append((offset + i, datetime(cell.year, cell.month, cell.day)))
    return date_cols


def _load_translations(wb: openpyxl.Workbook) -> dict[str, str | None]:
    """Italian test_name -> German translation, when the German sheet is usable.

    Stools.xlsx's German sheet is a stub (cells literally read "Loading...")
    so it is skipped; Blood-Urine's German sheet is complete and supplies a
    real translation used only as raw_payload provenance, never for mapping.
    """
    if _TRANSLATION_SHEET not in wb.sheetnames:
        return {}
    sheet = wb[_TRANSLATION_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}
    it_rows = wb[_PRIMARY_SHEET].iter_rows(values_only=True)
    it_names = [r[0] for r in it_rows][1:]
    translations: dict[str, str | None] = {}
    for it_name, de_row in zip(it_names, rows[1:], strict=False):
        if not it_name:
            continue
        de_name = de_row[0] if de_row else None
        if de_name is None or str(de_name).strip().lower() == "loading...":
            continue
        translations[str(it_name).strip()] = str(de_name).strip()
    return translations


def _parse_cell_value(raw_value: Any) -> tuple[float | None, str | None]:
    if isinstance(raw_value, int | float):
        return float(raw_value), None
    text = str(raw_value).strip()
    if not text:
        return None, None

    # Comparator-prefixed numerics ("<0.05", ">60") stay numeric like lab_pdf
    # does -- the qualifier itself is not represented, only the number.
    match = re.fullmatch(r"[<>]?\s*(\d+(?:[.,]\d+)?)", text)
    if match is not None:
        return float(match.group(1).replace(",", ".")), None

    # Genuinely qualitative ("negativ", "kein Material") -- never coerced.
    return None, text


def _normalize_unit(raw_unit: Any) -> str:
    # Sheet rows for enum/qualitative tests (e.g. "Hepatitis C Virus Ak") carry
    # no unit cell at all; Measurement.unit is non-optional, so an absent unit
    # normalizes to "" rather than inventing one.
    if raw_unit is None:
        return ""
    return str(raw_unit).strip()


def _measurement_uuid(
    metric_id: str, ts_utc: datetime, value_num: float | None, value_text: str | None
) -> str:
    value_repr = value_num if value_num is not None else value_text
    payload = f"{SOURCE_KIND}:{metric_id}:{ts_utc.isoformat()}:{value_repr}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", normalized.lower()).strip("-") or "field"


__all__ = ["LabXlsxParser"]
